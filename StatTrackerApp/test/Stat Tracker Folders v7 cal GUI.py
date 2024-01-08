# Stat Tracker Folders v3.py

# Standard library imports
import datetime
import importlib
import os
import subprocess
import threading
import time
import tkinter as tk
from tkinter import messagebox, ttk, font

# Third-party library imports
from PIL import Image, ImageTk
import pandas as pd
from openpyxl import load_workbook
from tkcalendar import Calendar
import webbrowser
import queue

# Import configuration from external file
from config import team_roster, event_codes

# Initialize global variables
event_log = []  # Stores the log of events for the current game
gui_update_queue = queue.Queue()  # Queue for GUI updates
stop_threads = False  # Flag to control thread execution

# ========================FUNCTIONS===========================================
def parse_date(date_str):
    try:
        return datetime.datetime.strptime(date_str, "%m/%d/%y").strftime("%m.%d.%y")
    except ValueError:
        raise ValueError("Invalid date format. Please use MM/DD/YY.")

def parse_time(time_str):
    try:
        return datetime.datetime.strptime(time_str, "%H:%M").strftime("%I:%M%p")
    except ValueError:
        raise ValueError("Invalid time format. Please use HH:MM in 24-hour format.")

def start_new_game_log():
    global game_info, event_log
    game_info = {}  # Initialize a dictionary to store game info

    try:
        game_info["date"] = parse_date(date_entry.get().strip())
        game_info["start_time"] = parse_time(start_time_entry.get().strip())
    except ValueError as e:
        messagebox.showerror("Input Error", str(e))
        return  # Exit the function if there's an error

    game_info["location"] = location_entry.get()
    game_info["opponent"] = opponent_entry.get()
    game_info["quarter"] = quarter_combobox.get().strip()

    game_info_response.config(
        text=f"{game_info['date']}\n{game_info['start_time']}\n{game_info['location']}\n{game_info['opponent']}\n{game_info['quarter']}",
        justify="left",
    )



def validate_player_number(player_number):
    try:
        player_number = int(player_number)
        if player_number not in team_roster:
            raise ValueError(f"Player number {player_number} is not valid.")
        return player_number
    except ValueError:
        raise ValueError("Player number must be an integer.")

def validate_event_code(event_code):
    if event_code not in event_codes:
        raise ValueError(f"Event code '{event_code}' is not valid.")

def add_event(video_time, player_number, event_code):
    try:
        player_number = validate_player_number(player_number)
        validate_event_code(event_code)

        player_name = ' '.join(team_roster[player_number][:2])
        event_description = event_codes[event_code]

        event_data = [
            game_info["date"],
            game_info["start_time"],
            game_info["location"],
            game_info["opponent"],
            game_info["quarter"],
            video_time,
            str(player_number),
            *team_roster[player_number][:2],
            event_description,
            event_code,
        ]
        event_log.append(event_data)
        event_log_text.insert(
            tk.END, f"{video_time} #{player_number} {player_name} {event_description}\n"
        )
        clear_event_entry()
        video_time_entry.focus()
        video_time_entry.select_range(0, tk.END)
    except ValueError as e:
        messagebox.showerror("Event Entry Error", str(e))


def handle_event_entry(event):
    video_time_input = video_time_entry.get()
    player_number = player_number_entry.get()
    event_code = event_code_entry.get()[:1]

    try:
        # Validate video time input format (e.g., "MM:SS")
        if (
            len(video_time_input) != 5
            or video_time_input[2] != ":"
            or not video_time_input[:2].isdigit()
            or not video_time_input[3:].isdigit()
        ):
            raise ValueError("Invalid video time format. Use 'MM:SS'.")

        # Validate player number
        if not player_number.isnumeric() or int(player_number) not in team_roster:
            raise ValueError("Invalid player number.")

        # Validate event code
        if event_code not in event_codes:
            raise ValueError("Invalid event code.")

        add_event(video_time_input, player_number, event_code)
    except ValueError as e:
        messagebox.showerror("Event Entry Error", str(e))


def clear_event_entry():
    # Assuming video_time_entry, player_number_entry, and event_code_entry are the entry widgets you want to clear
    video_time_entry.delete(0, tk.END)
    player_number_entry.delete(0, tk.END)
    event_code_entry.delete(0, tk.END)


def clear_entry_widgets(*entry_widgets):
    for entry_widget in entry_widgets:
        entry_widget.delete(0, tk.END)


def clear_all_data():
    clear_entry_widgets(
        date_entry,
        start_time_entry,
        location_entry,
        opponent_entry,
        quarter_entry,
        video_time_entry,
        player_number_entry,
        event_code_entry,
    )
    global game_info, event_log
    game_info = {}
    event_log = []
    event_log_text.delete("1.0", tk.END)


def load_workbook_template(path):
    if not os.path.exists(path):
        raise FileNotFoundError("Excel template file not found.")
    return load_workbook(path)


def fill_sheet_with_data(sheet, data):
    start_row = sheet.max_row + 1
    for entry in data:
        for col_idx, value in enumerate(entry, start=1):
            sheet.cell(row=start_row, column=col_idx, value=value)
        start_row += 1

def save_and_open_workbook(workbook, path):
    workbook.save(path)
    try:
        webbrowser.open(path)
    except Exception as e:
        raise Exception(f"Unable to open the Excel file: {e}")

def export_game_data_to_excel():
    try:
        opponent_without_spaces = game_info["opponent"].replace(" ", "_")
        desktop_path = os.path.expanduser("~/Desktop/Stat Tracker App")
        template_path = "../data/CSV to XL MASTER v3.xlsx"
        workbook = load_workbook_template(template_path)

        if "Raw Data" not in workbook.sheetnames:
            export_status_label.config(
                text="Error: 'Raw Data' sheet not found in the template."
            )
            return

        sheet = workbook["Raw Data"]
        fill_sheet_with_data(sheet, event_log)

        excel_filename = os.path.join(
            "../output", f"{game_info['date']}_{opponent_without_spaces}.xlsx"
        )
        save_and_open_workbook(workbook, excel_filename)
        export_status_label.config(
            text=f"Game data exported to Excel at {excel_filename}"
        )

    except FileNotFoundError as e:
        export_status_label.config(text=str(e))
    except PermissionError:
        export_status_label.config(text="Error: Permission denied for file operations.")
    except Exception as e:
        export_status_label.config(text=str(e))

def simulate_enter_key():
    event = tk.Event()
    event.keysym = "Return"
    event.event_type = "KeyPress"
    root.event_generate("<KeyPress-Return>", when="tail")
    root.event_generate("<KeyRelease-Return>", when="tail")
    handle_event_entry(event)
    toggle_button()


def toggle_button():
    global is_play_mode
    print("Button clicked. Current mode:", "Play" if is_play_mode else "Capture")
    if is_play_mode:
        on_play_click()
        button.config(text="STOP")
        is_play_mode = False
    else:
        on_capture_click()
        button.config(text="PLAY")
        is_play_mode = True
    root.update_idletasks()
    print("New mode:", "Play" if is_play_mode else "Capture")


def on_capture_click():
    try:
        # Capture timecode from QuickTime
        script_path = "../scripts/CaptureQTTimecode.scpt"
        result = subprocess.run(
            ["osascript", script_path], capture_output=True, text=True
        )

        # Check if subprocess run was successful
        if result.returncode != 0:
            raise RuntimeError("Failed to execute AppleScript.")

        captured_timecode = result.stdout.strip()  # Captured timecode in MMSS format
        print(f"Captured Timecode: {captured_timecode}")  # Debugging

        # Validate the format of captured timecode
        if len(captured_timecode) != 4 or not captured_timecode.isdigit():
            raise ValueError("Captured timecode format is incorrect.")

        total_seconds = int(captured_timecode[:2]) * 60 + int(captured_timecode[2:])
        adjustment_value = float(time_adjustment_spinbox.get())
        adjusted_timecode_seconds = max(total_seconds + adjustment_value, 0)
        adjusted_minutes = int(adjusted_timecode_seconds // 60)
        adjusted_seconds = int(adjusted_timecode_seconds % 60)
        adjusted_timecode = f"{adjusted_minutes:02}:{adjusted_seconds:02}"

        video_time_entry.delete(0, tk.END)  # Clear existing content
        video_time_entry.insert(0, adjusted_timecode)  # Insert adjusted timecode
        print("Capture script executed successfully")
    except subprocess.SubprocessError:
        print("Error: Unable to run the AppleScript.")
    except ValueError as ve:
        print(f"Value Error: {ve}")
    except RuntimeError as re:
        print(f"Runtime Error: {re}")
    except Exception as e:
        print(f"General Error executing capture script: {e}")


def on_play_click():
    try:
        script_path = "../scripts/Play Normal Speed.scpt"
        # Run the AppleScript
        result = subprocess.run(["osascript", script_path], check=True)

        # Check if the script executed successfully
        if result.returncode != 0:
            raise subprocess.SubprocessError(
                "AppleScript did not execute successfully."
            )

        print("Play script executed successfully")
    except subprocess.SubprocessError as se:
        print(f"Subprocess Error: {se}")
    except FileNotFoundError:
        print("Error: AppleScript file not found.")
    except PermissionError:
        print("Error: Permission denied for running the AppleScript.")
    except Exception as e:
        print(f"General Error executing play script: {e}")


def get_quicktime_timecode():
    """Get the current timecode from QuickTime Player using AppleScript."""
    script = """
        tell application "QuickTime Player"
            set timeCode to (current time of document 1) as string
        end tell
    """
    try:
        result = subprocess.run(
            ["osascript", "-e", script], capture_output=True, text=True, check=True
        )

        # Check if subprocess run was successful
        if result.returncode != 0:
            raise subprocess.SubprocessError(
                "AppleScript did not execute successfully."
            )

        return result.stdout.strip()
    except subprocess.SubprocessError as se:
        print(f"Subprocess Error: {se}")
        return None  # or an appropriate default/fallback value
    except FileNotFoundError:
        print("Error: AppleScript file not found.")
        return None
    except PermissionError:
        print("Error: Permission denied for running the AppleScript.")
        return None
    except Exception as e:
        print(f"General Error executing AppleScript: {e}")
        return None


def update_timecode():
    global stop_threads

    while not stop_threads:
        try:
            timecode_str = get_quicktime_timecode()

            # Validate the returned timecode string
            if timecode_str is None or not timecode_str.replace(".", "", 1).isdigit():
                raise ValueError("Invalid timecode format")

            # Convert the timecode to seconds as a float
            timecode_seconds = float(timecode_str)

            # Calculate minutes, seconds, and hundredths of a second
            minutes = int(timecode_seconds // 60)
            seconds = int(timecode_seconds % 60)
            hundredths = int((timecode_seconds - int(timecode_seconds)) * 100)

            # Format as 'minutes:seconds.hundredths'
            formatted_timecode = f"{minutes:02}:{seconds:02}:{hundredths:02}"

            # Put the update function in the queue
            gui_update_queue.put(lambda: timecode_label.config(text=formatted_timecode))

        except ValueError as ve:
            gui_update_queue.put(lambda: timecode_label.config(text="Invalid Timecode"))
            print(f"Value Error: {ve}")
        except Exception as e:
            gui_update_queue.put(lambda: timecode_label.config(text="Error"))
            print(f"Error: {e}")

        # Short sleep to avoid high CPU usage
        time.sleep(0.05)

def update_label(formatted_timecode):
    """Function to update the timecode label."""
    timecode_label.config(text=formatted_timecode)


def create_image_widget(parent, path, size=(120, 120)):
    try:
        img = Image.open(path)
        img = img.resize(size)
        tk_img = ImageTk.PhotoImage(img)
        img_label = tk.Label(parent, image=tk_img)
        img_label.image = tk_img  # Keep a reference to avoid garbage collection
        return img_label
    except FileNotFoundError:
        print(f"Error: The image file {path} was not found.")
        return None
    except IOError:
        print(f"Error: There was an issue opening the image file {path}.")
        return None
    except Exception as e:
        print(f"Unexpected error: {e}")
        return None


def on_combobox_click(
    event,
):  # Generate a virtual event for pressing the dropdown arrow
    event.widget.event_generate("<Down>")


def confirm_export():
    response = messagebox.askyesno(
        "Confirmation", "Do you want to export the game data?"
    )
    if response:
        export_game_data_to_excel()  # Call the original function if confirmed


def open_calendar(event, entry_widget):
    if not hasattr(open_calendar, "is_open") or not open_calendar.is_open:
        open_calendar.is_open = True

        def on_date_select(event=None):
            chosen_date = cal.get_date()
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, chosen_date)
            top.destroy()
            open_calendar.is_open = False

        top = tk.Toplevel(root)
        top.transient(root)  # Set top window as transient to root
        cal = Calendar(top, selectmode="day")
        cal.pack(fill="both", expand=True)

        # Bind the calendar's select event
        cal.bind("<<CalendarSelected>>", on_date_select)





def confirm_quit():
    global stop_threads
    response = messagebox.askyesno("Confirmation", "Are you sure you want to quit?")
    if response:
        stop_threads = True  # Signal threads to stop
        root.destroy()  # Close the main window



def handle_player_selection(event):  # Processes player selection from listbox.
    selected_index = player_listbox.curselection()
    if selected_index:
        selected_player = player_listbox.get(selected_index[0])
        # Extract the player number from the selected text (assuming the format "  {number}:  {first_name} {last_name}")
        player_number = selected_player.split(":")[0].strip()
        player_number_entry.delete(0, tk.END)  # Clear the player number entry
        player_number_entry.insert(0, player_number)  # Set the player number entry
        player_number_entry.focus()  # Set focus to the player number entry


def handle_event_code_selection(event):  # Handles event code selection from listbox.
    selected_index = event_code_listbox.curselection()
    if selected_index:
        selected_event_code = event_code_listbox.get(selected_index[0])
        # Extract the event code from the selected text (assuming the format "  {code}:  {description}")
        event_code = selected_event_code.split(":")[0].strip()
        event_code_entry.delete(0, tk.END)  # Clear the event code entry
        event_code_entry.insert(0, event_code)  # Set the event code entry
        event_code_entry.focus()  # Set focus to the event code entry


def format_time(sv):
    value = sv.get().replace(":", "")
    if len(value) == 4 and value.isdigit():
        # Only update if the value does not already have a colon
        if ":" not in sv.get():
            formatted = value[:2] + ":" + value[2:]
            sv.set(formatted)
    elif len(value) > 4:
        # Trim the value to 4 digits if more are entered
        sv.set(value[:4])


def select_all(entry):
    # Select all text in the entry widget
    entry.select_range(0, tk.END)
    # Move the cursor to the end
    entry.icursor(tk.END)


# Function to open the TeamRosterEditor window
def open_team_roster_editor():
    import TeamRosterEditor  # Import the TeamRosterEditor script
    importlib.reload(TeamRosterEditor)  # Reload the module to apply any changes


def process_gui_updates():
    while not gui_update_queue.empty():
        update_action = gui_update_queue.get()
        update_action()
    root.after(100, process_gui_updates)  # Schedule to check for updates every 100ms

# ================================ GUI SETUP ==============================
root = tk.Tk()
root.title("Stat Tracker")
root.configure(bg="darkorchid4")

# Font styles and style configuration
large_font = font.Font(family="Helvetica", size=20)
medium_font = font.Font(family="Arial", size=14)
small_font = font.Font(family="Helvetica", size=12)
style = ttk.Style()
style.configure("TButton", font=medium_font)

# Entry width for input fields
entry_width = 10

# ================================ LEFT FRAME SETUP ==============================
frame_left = ttk.Frame(root, borderwidth=2, relief="solid")
frame_left.grid(column=0, row=0, padx=10, pady=10, sticky="n")

# Game Info Frame
game_frame = ttk.Frame(frame_left)
game_frame.grid(column=0, row=0, pady=10, sticky="n")
game_frame.columnconfigure(1, weight=1)


game_info_label = tk.Label(
    game_frame, text="G A M E   I N F O", bg="#333333", fg="#666666"
)
game_info_label.grid(column=0, row=0, sticky="nw")

date_label = ttk.Label(game_frame, text="Date:")
date_label.grid(column=0, row=1, sticky="w")
date_entry = ttk.Entry(game_frame, width=entry_width)
date_entry.grid(column=1, row=1)
date_entry.bind("<Button-1>", lambda event: open_calendar(event, date_entry))

start_time_label = ttk.Label(game_frame, text="Game Start:")
start_time_label.grid(column=0, row=2, sticky="w")
start_time_entry = ttk.Entry(game_frame, width=entry_width)
start_time_entry.grid(column=1, row=2)
start_time_entry.bind("<FocusIn>", lambda event: select_all(start_time_entry))
start_time_entry.bind("<Return>", lambda event: location_entry.focus())

location_label = ttk.Label(game_frame, text="Venue:")
location_label.grid(column=0, row=3, sticky="w")
location_entry = ttk.Entry(game_frame, width=entry_width)
location_entry.grid(column=1, row=3)
location_entry.bind("<FocusIn>", lambda event: select_all(location_entry))
location_entry.bind("<Return>", lambda event: opponent_entry.focus())

opponent_label = ttk.Label(game_frame, text="Opponent:")
opponent_entry = ttk.Entry(game_frame, width=entry_width)
opponent_label.grid(column=0, row=4, sticky="w")
opponent_entry.grid(column=1, row=4)
opponent_entry.bind("<FocusIn>", lambda event: select_all(opponent_entry))
opponent_entry.bind("<Return>", lambda event: quarter_entry.focus())

quarter_label = ttk.Label(game_frame, text="Timeframe:")
quarter_label.grid(column=0, row=5, sticky="w")
quarter_entry = ttk.Entry(game_frame, width=entry_width)

# creates a combobox
options = [
    " Full Game",
    " 1st Quarter",
    " 2nd Quarter",
    " 3rd Quarter",
    " 4th Quarter",
    " Overtime",
    " First Half",
    " Second Half",
    " 5th Period",
]
quarter_combobox = ttk.Combobox(game_frame, values=options, width=9, state="readonly")
quarter_combobox.grid(column=1, row=5)
quarter_combobox.set("      SELECT")
quarter_entry.bind("<Return>", lambda event: start_new_game_log())
quarter_combobox.bind("<<ComboboxSelected>>", lambda event: start_new_game_log())
quarter_combobox.bind("<Button-1>", on_combobox_click)

# Default values for Entry fields (comment out if not needed)
date_entry.insert(0, "1/1/11")
start_time_entry.insert(0, "00:00")
location_entry.insert(0, "United Center")
opponent_entry.insert(0, "Bulls")
# quarter_entry.insert(0, "1")

# Create a horizontal separator for frame_left
separator1 = ttk.Separator(frame_left, orient="horizontal")
separator1.grid(row=1, columnspan=1, sticky="ew")
separator2 = ttk.Separator(frame_left, orient="horizontal")
separator2.grid(row=4, columnspan=1, pady=10, sticky="ew")
separator3 = ttk.Separator(frame_left, orient="horizontal")
separator3.grid(row=17, columnspan=1, pady=10, sticky="ew")



# ==========================================CONTROL SECTION================================================

# Timecode Label
timecode_label = tk.Label(
    frame_left,
    text="00:00:00",
    relief="raised",
    font=large_font,
    bg="black",
    fg="yellow green",
    height=1,
    width=7,
)
timecode_label.grid(column=0, row=2, padx=5, pady=10, sticky="w")


# Create a Spinbox widget for time adjustment with a specific width
time_adjustment_spinbox = ttk.Spinbox(
    frame_left,
    from_=-5.0,
    to=0.0,
    format="%.1f",
    font=small_font,
    increment=0.5,
    width=3,
)
time_adjustment_spinbox.grid(column=0, row=3, padx=5, sticky="w")
time_adjustment_spinbox.set("0.0")

# Play Button
is_play_mode = True  # Initially in 'Play' mode
button = ttk.Button(
    frame_left, text="PLAY", command=toggle_button, width=7, style="TButton"
)
button.grid(column=0, row=2, columnspan=2, padx=10, sticky="e")
# root.bind("<Key-space>", lambda event: toggle_button())

# Enter Button
enter_button = ttk.Button(
    frame_left, text="⏎", command=simulate_enter_key, width=7, style="TButton"
)
enter_button.grid(column=0, row=3, padx=10, pady=0, sticky="e")
# root.bind("<Key-Return>", lambda event: simulate_enter_key())

# Create an export Button with confirmation
export_button = ttk.Button(
    frame_left, text="Export Game Data", command=confirm_export, width=18
)
export_button.grid(column=0, row=20, columnspan=1, pady=0)

# Create a Quit Data Button with confirmation
quit_button = ttk.Button(frame_left, text="Quit", command=confirm_quit, width=18)
quit_button.grid(column=0, row=21, columnspan=1, pady=0)

# ========================================== EVENT ENTRY FRAME===============================

# Event Entry Frame
event_frame = ttk.Frame(frame_left)
event_frame.grid(column=0, row=5, padx=10, pady=10, sticky="w")
event_frame.columnconfigure(1, weight=1)

# Add a blank box with a label
blank_label = tk.Label(event_frame, text="E V E N T", bg="#333333", fg="#666666")
blank_label.grid(column=0, row=1, sticky="w")

video_time_label = ttk.Label(event_frame, text="Video Time:")
video_time_label.grid(column=0, row=2, sticky="w")
sv_video_time = tk.StringVar()
sv_video_time.trace_add(
    "write", lambda name, index, mode, sv=sv_video_time: format_time(sv)
)
video_time_entry = ttk.Entry(
    event_frame, width=10, justify="right", textvariable=sv_video_time
)
video_time_entry.grid(column=1, row=2)
video_time_entry.bind("<Return>", lambda event: player_number_entry.focus())

player_number_label = ttk.Label(event_frame, text="Player Number:")
player_number_label.grid(column=0, row=3, sticky="w")
player_number_entry = ttk.Entry(event_frame, width=entry_width, justify="right")
player_number_entry.grid(column=1, row=3)
player_number_entry.bind("<Return>", lambda event: event_code_entry.focus())

event_code_label = ttk.Label(event_frame, text="Event Code:")
event_code_label.grid(column=0, row=4, sticky="w")
event_code_entry = ttk.Entry(event_frame, width=entry_width, justify="right")
event_code_entry.grid(column=1, row=4)
event_code_entry.bind("<Return>", lambda event: handle_event_entry(event))

# Relative path to the image file in the resources or assets folder
image_path = "../resources/RMHS LogoShadow.png"
logo_widget = create_image_widget(frame_left, image_path)

if logo_widget is not None:
    logo_widget.grid(column=0, row=18, pady=10)
else:
    # Handle the error, such as logging or showing a message
    pass  # Placeholder for handling the error

# RMHS Label
rmhs_label = tk.Label(frame_left, text="RMHS Feeder Team", bg="#333333", fg="#666666")
rmhs_label.grid(column=0, row=19, pady=0)

# Add a blank box with a label
blank_label = tk.Label(frame_left, text="")
blank_label.grid(column=0, row=22, sticky="w")

# ================================ CENTER FRAME SETUP ==============================
frame_center = ttk.Frame(root, borderwidth=2, relief="solid")
frame_center.grid(column=1, row=0, padx=0, pady=10, sticky="n")

# Player info listbox
player_info_label = tk.Label(
    frame_center, text="P L A Y E R   I N F O", bg="#333333", fg="#666666"
)
player_info_label.grid(column=0, row=1, padx=10, pady=10, sticky="w")
player_listbox = tk.Listbox(frame_center, height=10, width=20)
player_listbox.grid(column=0, row=2, padx=10, ipady=5, pady=0)
for number, (first_name, last_name) in team_roster.items():
    player_listbox.insert(tk.END, f"  {number}:  {first_name} {last_name}")
player_listbox.bind("<<ListboxSelect>>", handle_player_selection)

# Create a button to open the TeamRosterEditor window
open_editor_button = ttk.Button(frame_center, text="edit", width="1", command=open_team_roster_editor)
open_editor_button.grid(column=0, row=1, padx=0, pady=10, sticky="e")

#create a horizontal separator for frame_center
separator4 = ttk.Separator(frame_center, orient="horizontal")
separator4.grid(row=3, columnspan=1, pady=10, sticky="ew")

# Event code listbox
event_code_label = tk.Label(
    frame_center, text="E V E N T   C O D E S", bg="#333333", fg="#666666"
)
event_code_label.grid(column=0, row=4, padx=10, pady=10, sticky="w")
event_code_listbox = tk.Listbox(frame_center, height=15, width=20)
event_code_listbox.grid(column=0, row=5, ipady=5, pady=0)

# Add a blank box with a label
blank_label = tk.Label(
    frame_center, text="Input:\nG A M E\nI N F O", bg="#333333", fg="#666666", height=6
)
blank_label.grid(column=0, row=6, pady=6, sticky="n")

for code, description in event_codes.items():
    event_code_listbox.insert(tk.END, f"  {code}:  {description}")
event_code_listbox.bind("<<ListboxSelect>>", handle_event_code_selection)

# Create the game info label
game_info_response = tk.Label(frame_center, bg="#333333", fg="#666666")
game_info_response.grid(column=0, row=6)

# ================================ RIGHT FRAME SETUP ==============================
frame_right = ttk.Frame(root, borderwidth=2, relief="solid")
frame_right.grid(column=2, row=0, padx=10, pady=10)

# Event Log Label
event_log_label = tk.Label(
    frame_right, text="E V E N T   L O G", bg="#333333", fg="#666666"
)
event_log_label.grid(column=0, row=0, pady=9, padx=10, sticky="w")
event_log_frame = ttk.Frame(frame_right)
event_log_frame.grid(column=0, row=1, columnspan=1)
event_log_frame.columnconfigure(0, weight=1)
event_log_frame.rowconfigure(0, weight=1)
custom_font = ("Courier", 14)
event_log_text = tk.Text(
    event_log_frame,
    height=43,
    width=47,
    font=custom_font,
    bg="black",
    fg="yellow green",
)
event_log_text.grid(row=0, column=0, padx=10, pady=0, sticky="nsew")

# Create a Scrollbar and set its command to the Text widget's yview
scrollbar = ttk.Scrollbar(
    event_log_frame, orient="vertical", command=event_log_text.yview
)
scrollbar.grid(row=0, column=1, sticky="ns")
event_log_text["yscrollcommand"] = scrollbar.set

# Create export status label within event log frame
export_status_label = ttk.Label(event_log_frame, text="", font=custom_font)
export_status_label.grid(row=1, column=0, padx=10, sticky="nsew")

# Starting the thread
threading.Thread(target=update_timecode, daemon=False).start()

# Start the process_gui_updates function
root.after(100, process_gui_updates)

# Start the main GUI loop
root.mainloop()
